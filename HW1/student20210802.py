#!/usr/bin/python3
from openpyxl import Workbook
from openpyxl import load_workbook 

read=load_workbook(filename='student.xlsx')
sheet_ranges=read['Sheet1']
wf=Workbook()
write=wf.active
write.title='Sheet1'

list1=[]
list2=[]
list3=[]
list4=[]
list5=[]

dic={}
dic2={}
dic3={}
aa=0
a0=0 
bb=0 
b0=0
cc=0
c0=0
f=0

for row in range(1,76):
    for column in range(1,9):
        a=sheet_ranges.cell(row=row, column=column).value
        write.cell(row=row, column=column, value=a)
        
for row in range(2,76):
    total=0
    total+=sheet_ranges.cell(row=row, column=3).value*0.3
    total+=sheet_ranges.cell(row=row, column=4).value*0.35
    total+=sheet_ranges.cell(row=row, column=5).value*0.34
    total+=sheet_ranges.cell(row=row, column=6).value*1
    write.cell(row=row, column=7, value=total)
    num=sheet_ranges.cell(row=row, column=7).value
    dic={row:total}        
    list1.append(dic)    

list2=sorted(list1, key=lambda item: list(item.values())[0], reverse=True)

i=0
j=1

while j<=len(list2):
    total=list(list2[i].values())[0]
    if total<40:
        dic2={list(list2[i].keys())[0]:'F'}
        list3.append(dic2)
        f+=1
        
    elif j<=len(list2)*0.15:
        dic2={list(list2[i].keys())[0]:'A+'}
        list3.append(dic2)      
        aa+=1
        
    elif j<len(list2)*0.30:
        dic2={list(list2[i].keys())[0]:'A0'}
        list3.append(dic2)
        a0+=1
        
    elif j<len(list2)*0.50:
        dic2={list(list2[i].keys())[0]:'B+'}
        list3.append(dic2)
        bb+=1
        
    elif j<len(list2)*0.70:
        dic2={list(list2[i].keys())[0]:'B0'}
        list3.append(dic2)   
        b0+=1
  
        
    elif j<len(list2)*1:
        dic2={list(list2[i].keys())[0]:'C0'}
        list3.append(dic2)
        c0+=1
        
    i+=1
    j+=1


if c0>1:    
    cc=c0//2

countCC=0   

i=0
j=1

while j<=len(list2):
    total=list(list2[i].values())[0]
    if total<40:
        dic3={list(list2[i].keys())[0]:'F'}
        list5.append(dic3)
        f+=1
        
    elif j<=len(list2)*0.15:
        dic3={list(list2[i].keys())[0]:'A+'}
        list5.append(dic3)      
        
    elif j<len(list2)*0.30:
        dic3={list(list2[i].keys())[0]:'A0'}
        list5.append(dic3)
        
    elif j<len(list2)*0.50:
        dic3={list(list2[i].keys())[0]:'B+'}
        list5.append(dic3)
        
    elif j<len(list2)*0.70:
        dic3={list(list2[i].keys())[0]:'B0'}
        list5.append(dic3)   
  
    elif j<len(list2)*0.85:
        if countCC<cc:
            dic3={list(list2[i].keys())[0]:'C+'}
            list5.append(dic3)   
            countCC+=1   
        else:
            dic3={list(list2[i].keys())[0]:'C0'}
            list5.append(dic3)
        
    elif j<len(list2)*1:
        dic3={list(list2[i].keys())[0]:'C0'}
        list5.append(dic3)
        
    i+=1
    j+=1


list4=sorted(list5, key=lambda item: list(item.keys())[0])


print(list3, "list3")
print(list4, "list4")
print(list5, "list5")
print(dic2,"dic2")
print(dic3,"dic3")


i=0
for row in range(2,76):
    write.cell(row=row, column=8, value=list(list4[i].values())[0])
    i+=1

wf.save('student.xlsx')













